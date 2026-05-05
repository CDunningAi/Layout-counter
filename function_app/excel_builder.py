"""
Excel workbook builder — produces a styled .xlsx from per-page detection results.

The workbook has:
  • One column per furniture category (YAML order).
  • One row per page (1-based page number).
  • A bold "Total" row at the bottom with column sums.
  • Bold, frozen header row.
  • Auto-sized columns.
"""

from __future__ import annotations

import io
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from logging_config import get_logger
from symbol_detector import PageResult

logger = get_logger(__name__)

_HEADER_FILL_HEX = "4472C4"  # Blue — matches Microsoft Office default table style
_TOTAL_FILL_HEX = "D9E1F2"   # Light blue


def build_workbook(page_results: List[PageResult], categories: List[str]) -> bytes:
    """
    Build a styled Excel workbook from *page_results* and return raw .xlsx bytes.

    Parameters
    ----------
    page_results : list[PageResult]
        Detection results, one entry per page (tiles already aggregated).
    categories : list[str]
        Ordered list of category names — determines column order.
    """
    if not page_results:
        logger.warning("build_workbook called with no page results — returning empty workbook")

    # Build a DataFrame with one row per page.
    rows = []
    for pr in page_results:
        row = {"Page": pr.page_num}
        for cat in categories:
            row[cat] = pr.counts.get(cat, 0)
        rows.append(row)

    df = pd.DataFrame(rows, columns=["Page"] + categories)
    df = df.sort_values("Page").reset_index(drop=True)

    # Append a Totals row.
    totals_row: dict = {"Page": "Total"}
    for cat in categories:
        totals_row[cat] = int(df[cat].sum())
    totals_df = pd.DataFrame([totals_row])
    df = pd.concat([df, totals_df], ignore_index=True)

    logger.info(
        "Building workbook: %d data rows, %d categories",
        len(page_results),
        len(categories),
    )

    # Write to an in-memory buffer using openpyxl engine.
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl", sheet_name="Layout Counts")

    # Re-open with openpyxl for styling.
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active

    # --- Style header row ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL_HEX)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Freeze the header row.
    ws.freeze_panes = "A2"

    # --- Style Total row (last row) ---
    total_row_idx = ws.max_row
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor=_TOTAL_FILL_HEX)
    for cell in ws[total_row_idx]:
        cell.font = total_font
        cell.fill = total_fill

    # --- Auto-size columns ---
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in column_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_length = max(max_length, cell_len)
            except Exception:  # noqa: BLE001
                pass
        # Add a small padding; openpyxl widths are approximately character widths.
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    # Save back to bytes.
    styled_buf = io.BytesIO()
    wb.save(styled_buf)
    xlsx_bytes = styled_buf.getvalue()

    logger.info("Workbook built: %d bytes", len(xlsx_bytes))
    return xlsx_bytes
